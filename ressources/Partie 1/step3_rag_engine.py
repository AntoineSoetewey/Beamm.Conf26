#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MÉTHODE 1 - STEP 3 : Moteur RAG + Boucle d'explication des variables
======================================================================

DEUX FONCTIONNALITÉS PRINCIPALES :

1. RECHERCHE SÉMANTIQUE SIMPLE
   - Recherche directe sur la collection unifiée ChromaDB
   - Aucune pondération par enquête (suppression de la stratégie hybride)
   - Score cosine brut utilisé pour le classement

2. BOUCLE D'EXPLICATION DES VARIABLES (nouvelle)
   - Parcourt chaque enquête et chaque variable
   - Génère des descriptions simplifiées par deux méthodes :
       a) Heuristique : nettoyage + structuration + détection du type
       b) LLM (Llama 3.2 via Ollama) : réécriture en langage accessible
   - Descriptions longues et détaillées (pas de résumés courts)
   - Langue originale conservée + version française ajoutée
   - Sauvegarde dans 1 fichier Excel par enquête (6 fichiers au total)
   - Checkpoint toutes les 50 variables (reprise en cas d'interruption)
   - Feuille "Résumé" dans chaque fichier Excel

ENQUÊTES SUPPORTÉES : EU-SILC, HFCS, EU-LFS, HBS, IPCAL, DEMOBEL

CRITIQUE : Doit utiliser le MÊME modèle que step2 !
- Modèle : intfloat/multilingual-e5-large (1024D)
- Préfixe : 'query:' pour les questions

Auteur: Carnot
Date: Mars 2026
"""

import json
import re
import time
import logging
from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime

try:
    import chromadb
    from sentence_transformers import SentenceTransformer
except ImportError as e:
    print(f"❌ Erreur: {e}")
    print("pip install chromadb sentence-transformers")
    exit(1)

try:
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"❌ Erreur: {e}")
    print("pip install pandas openpyxl")
    exit(1)

try:
    import ollama
    OLLAMA_AVAILABLE = True
except ImportError:
    OLLAMA_AVAILABLE = False

# ─── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s',
    handlers=[
        logging.FileHandler('step3_explainer.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
# CLASSE PRINCIPALE RAG (RECHERCHE SIMPLE — SANS PONDÉRATION)
# ══════════════════════════════════════════════════════════════════════════════

class MultiSurveyRAG:
    """
    Système RAG pour recherche sémantique avec E5-Large.

    VERSION SIMPLIFIÉE :
    - Recherche directe sur la collection unifiée
    - Score cosine brut, aucune pondération par enquête
    - Compatible avec métrique cosine (configurée dans step2)
    """

    def __init__(self, chroma_path: Path, variables_path: Path = None):
        self.chroma_path = Path(chroma_path)

        if not self.chroma_path.exists():
            raise FileNotFoundError(f"Index ChromaDB introuvable: {self.chroma_path}")

        # Charger E5-Large (même modèle que step2)
        print("\n📦 Chargement du modèle E5-Large...")
        self.model = SentenceTransformer('intfloat/multilingual-e5-large')
        print(f"   ✅ Modèle chargé (1024D)")

        # ChromaDB
        self.client = chromadb.PersistentClient(path=str(self.chroma_path))

        # Collections
        self.collections = {
            'unified': self.client.get_collection('unified_variables'),
            'eu_silc': self.client.get_collection('eu_silc_variables'),
            'hfcs'   : self.client.get_collection('hfcs_variables'),
            'eu_lfs' : self.client.get_collection('eu_lfs_variables'),
            'hbs'    : self.client.get_collection('hbs_variables'),
            'ipcal'  : self.client.get_collection('ipcal_variables'),
            'demobel': self.client.get_collection('demobel_variables'),
        }

        # Vérifier métrique
        collection_metadata = self.collections['unified'].metadata
        self.distance_metric = collection_metadata.get('hnsw:space', 'l2')

        if self.distance_metric == 'cosine':
            print(f"   📏 Métrique détectée: COSINE (standard NLP) ✅")
        else:
            print(f"   ⚠️  Métrique détectée: {self.distance_metric.upper()}")

        # Variables complètes
        self.variables_dict = {}
        self.variables_by_code = {}
        self.variables_by_survey: Dict[str, List[Dict]] = {}

        if variables_path and Path(variables_path).exists():
            with open(variables_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                for v in data['variables']:
                    self.variables_dict[v['variable_id']] = v
                    self.variables_by_code[v['code']] = v
                    survey = v.get('survey', 'UNKNOWN')
                    self.variables_by_survey.setdefault(survey, []).append(v)

            print(f"   📊 {len(self.variables_dict)} variables chargées")
            for survey, vars_list in self.variables_by_survey.items():
                print(f"      • {survey}: {len(vars_list)} variables")

    # ──────────────────────────────────────────────────────────────────────────
    # RECHERCHE SÉMANTIQUE SIMPLE (sans pondération)
    # ──────────────────────────────────────────────────────────────────────────

    def encode_query(self, question: str):
        """Encode une question avec E5-Large + préfixe 'query:'."""
        prefixed = f"query: {question}"
        embedding = self.model.encode(
            prefixed,
            normalize_embeddings=True,
            convert_to_numpy=True
        )
        return embedding.tolist()

    def _convert_distance_to_score(self, distance: float) -> float:
        """Convertit la distance cosine ChromaDB en score de similarité [0, 1]."""
        if self.distance_metric == 'cosine':
            return 1.0 - (distance / 2.0)
        return 1.0 / (1.0 + distance)

    def search_by_question(self, question: str, top_k: int = 5) -> List[Dict]:
        """
        Recherche sémantique simple sur la collection unifiée.
        Aucune pondération — score cosine brut utilisé directement.
        """
        collection = self.collections['unified']
        query_embedding = self.encode_query(question)

        results = collection.query(
            query_embeddings=[query_embedding],
            n_results=min(top_k, collection.count())
        )

        formatted = []
        for i in range(len(results['ids'][0])):
            distance = results['distances'][0][i]
            score    = self._convert_distance_to_score(distance)
            formatted.append({
                'variable_id': results['ids'][0][i],
                'distance'   : distance,
                'score'      : score,
                'metadata'   : results['metadatas'][0][i],
            })

        formatted.sort(key=lambda x: x['score'], reverse=True)
        return formatted[:top_k]

    def search_by_code(self, variable_code: str) -> Dict:
        """Recherche par code de variable (lookup direct)."""
        # Recherche insensible à la casse
        for key, var in self.variables_by_code.items():
            if key.upper() == variable_code.upper() or key.lower() == variable_code.lower():
                return {
                    'found'      : True,
                    'code'       : variable_code,
                    'survey'     : var['survey'],
                    'name'       : var['name_en'],
                    'category'   : var.get('category', 'N/A'),
                    'description': var.get('description_long',
                                   var.get('description_short', 'N/A'))
                }
        return {
            'found'  : False,
            'code'   : variable_code,
            'message': f"Variable '{variable_code}' non trouvée"
        }

    def display_results(self, results: List[Dict], top_k: int = 5):
        """Affiche les résultats avec score cosine brut."""
        for i, res in enumerate(results[:top_k], 1):
            meta        = res['metadata']
            code        = meta['code']
            survey      = meta['survey']
            score       = res['score']
            distance    = res['distance']

            if code in self.variables_by_code:
                var_info = self.variables_by_code[code]
                description = var_info.get('description_long',
                              var_info.get('description_short',
                              var_info.get('name_en', 'N/A')))
            else:
                description = meta.get('name_en', 'N/A')

            print(f"\n{i}. {code} ({survey})")
            print(f"   {description[:120]}")
            print(f"   Similarité: {score:.1%} | Distance: {distance:.3f}")


# ══════════════════════════════════════════════════════════════════════════════
# MOTEUR D'EXPLICATION DES VARIABLES
# ══════════════════════════════════════════════════════════════════════════════

class VariableExplainer:
    """
    Génère des descriptions simplifiées pour chaque variable de chaque enquête.

    MÉTHODE 1 — Heuristique :
        - Nettoyage des codes internes ([NL], VAK, --info--, |)
        - Détection de la langue (NL/FR/EN)
        - Inférence du type de variable (monétaire, catégorielle, date...)
        - Structuration en description courte et longue
        - Traduction/reformulation en français simple

    MÉTHODE 2 — LLM (Llama 3.2 via Ollama) :
        - Réécriture complète en langage accessible (b-a ba)
        - Descriptions longues et détaillées (5-8 phrases)
        - Exemple concret d'utilisation
        - Contexte et relation avec d'autres variables
        - Version originale + version française

    SORTIE : 1 fichier Excel par enquête avec feuille Résumé
    """

    # ── Indicateurs de type de variable ───────────────────────────────────────
    _TYPE_PATTERNS = {
        'monetaire'   : r'(income|revenue|revenu|inkomen|montant|amount|wage|salaire|'
                        r'loon|benefit|allocation|pension|loyer|rent|cadastr|valeur|value)',
        'pourcentage' : r'(rate|taux|percentage|pourcentage|ratio|proportion|percent)',
        'date'        : r'(year|année|jaar|date|month|mois|maand|quarter|trimestre)',
        'comptage'    : r'(number|nombre|aantal|count|nb_|n_|nbre|hours|heures|uren|'
                        r'weeks|semaines|weken|months|mois)',
        'categorielle': r'(type|status|statut|categor|categorie|niveau|level|code|'
                        r'indicator|indicateur|flag|yes|no|oui|non|boolean)',
        'geographique': r'(region|commune|province|country|pays|land|zip|postal|'
                        r'municipality|gemeente|arrondissement)',
    }

    # ── Patterns de nettoyage ──────────────────────────────────────────────────
    _CLEAN_PATTERNS = [
        (r'\[NL\]\s*',      ''),
        (r'\[FR\]\s*',      ''),
        (r'\[EN\]\s*',      ''),
        (r'\|',             ' — '),
        (r'VAK\s+',         ''),
        (r'--info--',       ''),
        (r'--.*?--',        ''),
        (r'\s{2,}',         ' '),
        (r'^\s*—\s*',       ''),
        (r'\s*—\s*$',       ''),
        (r'Code déclaration:\s*[\d.]+', ''),
        (r'Statut:\s*\w+',  ''),
    ]

    # ── Indicateurs de langue ─────────────────────────────────────────────────
    _LANG_NL = {'de', 'het', 'een', 'van', 'in', 'op', 'met', 'voor', 'zijn',
                'worden', 'heeft', 'hebben', 'bij', 'aan', 'uit', 'als', 'ook',
                'berekening', 'belasting', 'inkomen', 'inkomsten', 'aantal',
                'bedrag', 'jaar', 'maand', 'persoon', 'huishouden'}
    _LANG_FR = {'le', 'la', 'les', 'de', 'du', 'des', 'un', 'une', 'et', 'ou',
                'dans', 'sur', 'par', 'pour', 'avec', 'est', 'sont', 'revenus',
                'montant', 'calcul', 'impôt', 'déclaration', 'personne', 'ménage',
                'année', 'mois', 'nombre', 'valeur', 'type', 'cadre'}

    def __init__(self, output_dir: Path, ollama_model: str = 'llama3.2',
                 checkpoint_every: int = 50):
        self.output_dir       = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.ollama_model     = ollama_model
        self.checkpoint_every = checkpoint_every

        # Vérifier Ollama
        self.llm_available = False
        if OLLAMA_AVAILABLE:
            try:
                client = ollama.Client()
                client.list()
                self.llm_available = True
                self.ollama_client = client
                log.info(f"✅ Ollama disponible (modèle: {ollama_model})")
            except Exception as e:
                log.warning(f"⚠️  Ollama non accessible: {e} → mode heuristique uniquement")
        else:
            log.warning("⚠️  Ollama non installé → mode heuristique uniquement")

    # ──────────────────────────────────────────────────────────────────────────
    # MÉTHODE 1 : HEURISTIQUE
    # ──────────────────────────────────────────────────────────────────────────

    def _clean_text(self, text: str) -> str:
        """Nettoie les codes internes et artefacts des descriptions."""
        if not text:
            return ''
        result = str(text)
        for pattern, replacement in self._CLEAN_PATTERNS:
            result = re.sub(pattern, replacement, result, flags=re.IGNORECASE)
        return result.strip()

    def _detect_language(self, text: str) -> str:
        """Détecte la langue dominante : NL, FR ou EN."""
        if not text:
            return 'EN'
        words = set(re.findall(r'\b\w+\b', text.lower()))
        score_nl = len(words & self._LANG_NL)
        score_fr = len(words & self._LANG_FR)
        if score_nl > score_fr and score_nl >= 2:
            return 'NL'
        if score_fr > score_nl and score_fr >= 2:
            return 'FR'
        return 'EN'

    def _detect_type(self, code: str, name: str, desc: str) -> str:
        """Infère le type de variable depuis son code, nom et description."""
        text = f"{code} {name} {desc}".lower()
        for type_name, pattern in self._TYPE_PATTERNS.items():
            if re.search(pattern, text, re.IGNORECASE):
                return type_name
        return 'autre'

    def _detect_level(self, code: str, survey: str, desc: str) -> str:
        """Infère le niveau d'observation (ménage, individu, fiscal)."""
        text = f"{code} {desc}".lower()
        if survey == 'IPCAL':
            return 'fiscal'
        if re.search(r'(household|ménage|huishouden|hh_|_h_)', text):
            return 'ménage'
        if re.search(r'(person|personne|persoon|individual|_p_|_i_)', text):
            return 'individu'
        return 'inconnu'

    def _quality_score(self, desc_short: str, desc_long: str) -> int:
        """Évalue la qualité de la description originale (1=pauvre, 3=riche)."""
        total = len(desc_short or '') + len(desc_long or '')
        if total > 300:
            return 3
        if total > 80:
            return 2
        return 1

    def _heuristic_short_fr(self, var: Dict) -> str:
        """
        Génère une description courte simplifiée en français.
        Minimum 2-3 phrases complètes et compréhensibles.
        """
        code    = var.get('code', '')
        survey  = var.get('survey', '')
        name    = self._clean_text(var.get('name_en', '') or var.get('name_fr', ''))
        cat     = var.get('category', '') or ''
        subcat  = var.get('subcategory', '') or ''
        vtype   = self._detect_type(code, name, var.get('description_short', ''))
        level   = self._detect_level(code, survey, var.get('description_short', ''))

        # Carte d'identité de la variable
        type_labels = {
            'monetaire'   : 'une valeur monétaire (en euros)',
            'pourcentage' : 'un taux ou pourcentage',
            'date'        : 'une date, une année ou une période',
            'comptage'    : 'un nombre ou une quantité',
            'categorielle': 'une catégorie ou un indicateur (oui/non, type...)',
            'geographique': 'une zone géographique ou une localisation',
            'autre'       : 'une donnée statistique',
        }
        level_labels = {
            'ménage'   : 'au niveau du ménage (logement)',
            'individu' : 'au niveau de la personne',
            'fiscal'   : 'au niveau de la déclaration fiscale',
            'inconnu'  : '',
        }

        type_label  = type_labels.get(vtype, 'une donnée statistique')
        level_label = level_labels.get(level, '')

        parts = []
        parts.append(
            f"La variable {code}, issue de l'enquête {survey}, représente "
            f"{type_label}{' ' + level_label if level_label else ''}."
        )
        if name and name not in ('0', 'nan', ''):
            parts.append(
                f"Elle est intitulée : « {name} »."
            )
        if cat and cat not in ('0', 'nan', ''):
            parts.append(
                f"Elle appartient à la catégorie « {cat} »"
                + (f", sous-catégorie « {subcat} »" if subcat and subcat not in ('0','nan','') else '')
                + "."
            )
        parts.append(
            f"Cette variable a été collectée dans le cadre de l'enquête {survey} "
            f"et peut être utilisée dans des analyses statistiques portant sur "
            f"{'les finances et la fiscalité des ménages belges' if survey == 'IPCAL' else 'les conditions de vie et les revenus des ménages'}."
        )

        return ' '.join(parts)

    def _heuristic_long_fr(self, var: Dict) -> str:
        """
        Génère une description longue et détaillée en français.
        Minimum 6-8 phrases avec contexte, usage et interprétation.
        """
        code     = var.get('code', '')
        survey   = var.get('survey', '')
        name     = self._clean_text(var.get('name_en', '') or var.get('name_fr', ''))
        desc_s   = self._clean_text(var.get('description_short', '') or '')
        desc_l   = self._clean_text(var.get('description_long', '') or '')
        cat      = var.get('category', '') or ''
        subcat   = var.get('subcategory', '') or ''
        ref_per  = var.get('reference_period', '') or ''
        vtype    = self._detect_type(code, name, desc_s)
        level    = self._detect_level(code, survey, desc_s)
        lang     = self._detect_language(f"{name} {desc_s}")

        # Description originale nettoyée (préserver si elle est riche)
        orig = desc_l if len(desc_l) > len(desc_s) else desc_s
        orig = orig[:500] if orig else ''

        # Descriptions du type de valeur attendu
        type_explanations = {
            'monetaire'   : ('une valeur exprimée en euros',
                             'des montants financiers comme des revenus, '
                             'des dépenses ou des patrimoniaux'),
            'pourcentage' : ('un taux ou pourcentage (entre 0 et 100)',
                             'des proportions ou des taux de variation'),
            'date'        : ("une information temporelle (année, mois, période)",
                             'des chronologies ou des durées'),
            'comptage'    : ("un nombre entier (heures, mois, personnes...)",
                             'des fréquences, des durées ou des effectifs'),
            'categorielle': ("une catégorie discrète (ex. : oui/non, type 1/2/3)",
                             'des classifications ou des indicateurs qualitatifs'),
            'geographique': ("un identifiant géographique (région, commune...)",
                             'des analyses territoriales'),
            'autre'       : ("une valeur statistique",
                             'des analyses multivariées'),
        }
        type_val, type_usage = type_explanations.get(vtype, ('une valeur', 'des analyses'))

        level_context = {
            'ménage'   : ("Elle est mesurée au niveau du ménage, c'est-à-dire "
                          "qu'elle caractérise l'ensemble des personnes vivant "
                          "sous le même toit."),
            'individu' : ("Elle est mesurée au niveau individuel, c'est-à-dire "
                          "qu'elle décrit une caractéristique propre à chaque "
                          "personne enquêtée."),
            'fiscal'   : ("Elle est issue d'une déclaration fiscale individuelle "
                          "ou conjointe, et reflète une situation fiscale précise "
                          "pour l'exercice concerné."),
            'inconnu'  : '',
        }

        survey_contexts = {
            'EU-SILC'  : ("L'enquête EU-SILC (Statistics on Income and Living "
                          "Conditions) est menée annuellement dans tous les États "
                          "membres de l'Union européenne. Elle vise à mesurer les "
                          "revenus, la pauvreté et les conditions de vie des ménages."),
            'HFCS'     : ("L'enquête HFCS (Household Finance and Consumption Survey) "
                          "est coordonnée par la Banque Centrale Européenne. Elle "
                          "recueille des données détaillées sur le patrimoine, "
                          "l'endettement et la consommation des ménages."),
            'EU-LFS'   : ("L'enquête EU-LFS (Labour Force Survey) est la principale "
                          "source de données sur le marché du travail en Europe. "
                          "Elle mesure l'emploi, le chômage et les conditions "
                          "de travail des personnes en âge de travailler."),
            'HBS'      : ("L'enquête HBS (Household Budget Survey) analyse en détail "
                          "les dépenses des ménages en biens et services. Elle permet "
                          "de calculer des indices de prix et d'étudier les habitudes "
                          "de consommation."),
            'IPCAL'    : ("IPCAL est la base de données fiscale belge regroupant "
                          "les informations issues des déclarations d'impôt des "
                          "contribuables belges. Elle couvre les revenus professionnels, "
                          "immobiliers, mobiliers et divers."),
            'DEMOBEL'  : ("DEMOBEL est un modèle de microsimulation belge qui intègre "
                          "des données socioéconomiques détaillées sur les ménages "
                          "belges. Il est utilisé pour simuler l'impact de politiques "
                          "fiscales et sociales."),
        }

        parts = []

        # 1. Identification
        parts.append(
            f"La variable {code} est une variable de l'enquête {survey}. "
            f"Elle enregistre {type_val}."
        )

        # 2. Intitulé
        if name and name not in ('0', 'nan', ''):
            if lang == 'NL':
                parts.append(
                    f"Son intitulé original en néerlandais est : « {name} ». "
                    f"En français, cela signifie approximativement : une mesure "
                    f"relative à {self._clean_text(name).lower()}."
                )
            elif lang == 'EN':
                parts.append(
                    f"Son intitulé original en anglais est : « {name} ». "
                    f"Cette dénomination indique qu'il s'agit d'une mesure "
                    f"portant sur {self._clean_text(name).lower()}."
                )
            else:
                parts.append(
                    f"Son intitulé est : « {name} »."
                )

        # 3. Description originale enrichie
        if orig and len(orig) > 20:
            parts.append(
                f"De manière plus précise, voici ce que mesure cette variable : "
                f"{orig}"
            )

        # 4. Catégorie
        if cat and cat not in ('0', 'nan', ''):
            parts.append(
                f"Dans la structure de l'enquête {survey}, cette variable appartient "
                f"à la catégorie « {cat} »"
                + (f", plus précisément à la sous-catégorie « {subcat} »"
                   if subcat and subcat not in ('0', 'nan', '') else '')
                + ". Cette classification permet aux chercheurs de regrouper "
                  "les variables par thématique lors des analyses."
            )

        # 5. Niveau d'observation
        lctx = level_context.get(level, '')
        if lctx:
            parts.append(lctx)

        # 6. Contexte de l'enquête
        sctx = survey_contexts.get(survey, '')
        if sctx:
            parts.append(sctx)

        # 7. Période de référence
        if ref_per and ref_per not in ('', 'nan'):
            parts.append(
                f"La période de référence pour cette variable est : {ref_per}."
            )

        # 8. Usage analytique
        parts.append(
            f"Dans la pratique, cette variable est typiquement utilisée pour "
            f"des analyses portant sur {type_usage}. "
            f"Elle peut être croisée avec d'autres variables de la même enquête "
            f"pour produire des indicateurs composites ou des profils sociodémographiques."
        )

        # 9. Conseil d'interprétation
        parts.append(
            f"Pour interpréter correctement cette variable, il est important de "
            f"tenir compte de ses valeurs manquantes éventuelles, de son unité "
            f"de mesure et du contexte dans lequel elle a été collectée. "
            f"Tout chercheur souhaitant l'utiliser devrait consulter le dictionnaire "
            f"des variables officiel de l'enquête {survey} pour obtenir la définition "
            f"complète et les codes de valeurs associés."
        )

        return ' '.join(parts)

    # ──────────────────────────────────────────────────────────────────────────
    # MÉTHODE 2 : LLM (Llama 3.2 via Ollama)
    # ──────────────────────────────────────────────────────────────────────────

    def _llm_describe(self, var: Dict, lang_target: str = 'FR') -> Dict:
        """
        Génère des descriptions simplifiées via Llama 3.2.
        Retourne un dict avec 'short' et 'long' en langue originale + français.
        """
        if not self.llm_available:
            return {'short_orig': '', 'long_orig': '', 'short_fr': '', 'long_fr': ''}

        code    = var.get('code', '')
        survey  = var.get('survey', '')
        name    = var.get('name_en', '') or var.get('name_fr', '') or ''
        desc_s  = var.get('description_short', '') or ''
        desc_l  = var.get('description_long', '') or ''
        cat     = var.get('category', '') or ''
        lang    = self._detect_language(f"{name} {desc_s}")

        lang_labels = {'NL': 'néerlandais', 'FR': 'français', 'EN': 'anglais'}
        lang_label  = lang_labels.get(lang, 'anglais')

        # ── Prompt pour description COURTE (langue originale) ────────────────
        prompt_short_orig = f"""Tu es un expert en statistiques sociales et économiques.

Voici une variable statistique issue de l'enquête {survey} :

- Code : {code}
- Nom : {name}
- Catégorie : {cat}
- Description courte originale : {desc_s[:300]}
- Description longue originale : {desc_l[:500]}

Ta tâche : rédiger une description COURTE et ACCESSIBLE de cette variable EN {lang_label.upper()}.
Cette description doit :
1. Expliquer clairement CE QUE MESURE cette variable (2-3 phrases complètes)
2. Préciser le TYPE de valeur attendu (montant en euros, nombre, catégorie, etc.)
3. Donner un EXEMPLE CONCRET simple (ex: "Si une personne gagne 2000€/mois de loyer, cette variable vaudra 2000")
4. Être compréhensible par quelqu'un sans formation en statistiques
5. Ne pas utiliser de jargon technique sans l'expliquer

Rédige UNIQUEMENT la description, sans introduction ni conclusion."""

        # ── Prompt pour description LONGUE (langue originale) ────────────────
        prompt_long_orig = f"""Tu es un expert en statistiques sociales et économiques.

Voici une variable statistique issue de l'enquête {survey} :

- Code : {code}
- Nom : {name}
- Catégorie : {cat}
- Description courte originale : {desc_s[:300]}
- Description longue originale : {desc_l[:600]}

Ta tâche : rédiger une description LONGUE, DÉTAILLÉE et ACCESSIBLE de cette variable EN {lang_label.upper()}.
Cette description doit OBLIGATOIREMENT contenir (minimum 8 phrases) :
1. Une présentation claire de ce que mesure la variable
2. Le contexte de l'enquête {survey} et pourquoi cette variable y figure
3. Le type de valeur (montant, pourcentage, catégorie, date...) et son unité
4. Un exemple concret et parlant (situation réelle d'un ménage ou d'une personne)
5. Comment cette variable peut être utilisée dans une analyse
6. Les précautions d'interprétation (valeurs manquantes, cas particuliers)
7. La relation possible avec d'autres variables similaires
8. Une conclusion sur l'importance de cette variable pour la recherche

Rédige UNIQUEMENT la description complète, sans introduction ni conclusion."""

        # ── Prompt pour description COURTE (français) ────────────────────────
        prompt_short_fr = f"""Tu es un expert en statistiques sociales et économiques.

Voici une variable statistique issue de l'enquête {survey} :

- Code : {code}
- Nom : {name}
- Description en {lang_label} : {desc_s[:300]}

Ta tâche : rédiger une description COURTE et ACCESSIBLE de cette variable EN FRANÇAIS.
Cette description doit :
1. Expliquer clairement CE QUE MESURE cette variable (2-3 phrases complètes)
2. Préciser le TYPE de valeur (montant en euros, nombre, catégorie, etc.)
3. Donner un EXEMPLE CONCRET simple
4. Être compréhensible par quelqu'un sans formation en statistiques

Rédige UNIQUEMENT la description en français, sans introduction ni conclusion."""

        # ── Prompt pour description LONGUE (français) ────────────────────────
        prompt_long_fr = f"""Tu es un expert en statistiques sociales et économiques.

Voici une variable statistique issue de l'enquête {survey} :

- Code : {code}
- Nom : {name}
- Description originale en {lang_label} : {desc_l[:500] or desc_s[:300]}

Ta tâche : rédiger une description LONGUE, DÉTAILLÉE et ACCESSIBLE EN FRANÇAIS.
Cette description doit OBLIGATOIREMENT contenir (minimum 8 phrases) :
1. Une présentation claire de ce que mesure la variable
2. Le contexte de l'enquête {survey}
3. Le type de valeur et son unité de mesure
4. Un exemple concret et parlant
5. Comment utiliser cette variable dans une analyse
6. Les précautions d'interprétation
7. Les relations avec d'autres variables
8. L'importance de cette variable pour la recherche

Rédige UNIQUEMENT la description complète en français, sans introduction ni conclusion."""

        results = {'short_orig': '', 'long_orig': '', 'short_fr': '', 'long_fr': ''}

        for key, prompt in [
            ('short_orig', prompt_short_orig),
            ('long_orig',  prompt_long_orig),
            ('short_fr',   prompt_short_fr),
            ('long_fr',    prompt_long_fr),
        ]:
            # Si la langue originale est déjà le français, copier
            if key == 'short_fr' and lang == 'FR' and results['short_orig']:
                results['short_fr'] = results['short_orig']
                continue
            if key == 'long_fr' and lang == 'FR' and results['long_orig']:
                results['long_fr'] = results['long_orig']
                continue

            try:
                response = self.ollama_client.chat(
                    model=self.ollama_model,
                    messages=[
                        {
                            'role'   : 'system',
                            'content': ('Tu es un expert en statistiques qui explique '
                                        'des variables de manière simple et accessible '
                                        'à un large public. Tes descriptions sont toujours '
                                        'longues, précises et compréhensibles.')
                        },
                        {'role': 'user', 'content': prompt}
                    ],
                    options={'temperature': 0.3, 'num_predict': 800}
                )
                results[key] = response['message']['content'].strip()
                time.sleep(0.1)  # éviter la surcharge d'Ollama
            except Exception as e:
                log.warning(f"   ⚠️  LLM erreur pour {code} ({key}): {e}")
                results[key] = ''

        return results

    # ──────────────────────────────────────────────────────────────────────────
    # GÉNÉRATION DES FICHIERS EXCEL
    # ──────────────────────────────────────────────────────────────────────────

    def _style_excel(self, ws, n_rows: int):
        """Applique la mise en forme au fichier Excel."""
        # Styles
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        header_fill  = PatternFill('solid', start_color='1F4E79')
        header_align = Alignment(horizontal='center', vertical='center',
                                 wrap_text=True)
        cell_font    = Font(name='Arial', size=10)
        wrap_align   = Alignment(vertical='top', wrap_text=True)
        center_align = Alignment(horizontal='center', vertical='top')
        alt_fill     = PatternFill('solid', start_color='EBF3FB')

        # En-têtes
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = border
        ws.row_dimensions[1].height = 32

        # Données
        center_cols = {1, 2, 3, 4, 5, 6}
        for row_idx in range(2, n_rows + 2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, cell in enumerate(ws[row_idx], 1):
                cell.font   = cell_font
                cell.border = border
                cell.alignment = center_align if col_idx in center_cols else wrap_align
                if fill:
                    cell.fill = fill
            ws.row_dimensions[row_idx].height = 60

        # Largeurs de colonnes
        col_widths = {
            'A': 16, 'B': 12, 'C': 14, 'D': 10, 'E': 10, 'F': 10,
            'G': 55, 'H': 55, 'I': 55, 'J': 55, 'K': 55, 'L': 55,
        }
        for col_letter, width in col_widths.items():
            if col_letter in [get_column_letter(i) for i in range(1, ws.max_column + 1)]:
                ws.column_dimensions[col_letter].width = width

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{n_rows + 1}"

    def _add_summary_sheet(self, wb: Workbook, survey: str, records: List[Dict],
                           n_llm_ok: int, n_llm_fail: int, elapsed: float):
        """Ajoute une feuille Résumé au fichier Excel."""
        ws = wb.create_sheet("Résumé", 0)

        # Styles
        title_font   = Font(name='Arial', bold=True, size=14, color='1F4E79')
        section_font = Font(name='Arial', bold=True, size=11, color='2E75B6')
        normal_font  = Font(name='Arial', size=10)
        header_fill  = PatternFill('solid', start_color='1F4E79')
        header_font  = Font(name='Arial', bold=True, color='FFFFFF', size=10)

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 40

        row = 1
        ws.cell(row, 1, f"Rapport d'explication — Enquête {survey}").font = title_font
        ws.cell(row, 1).fill = PatternFill('solid', start_color='EBF3FB')
        row += 1
        ws.cell(row, 1, f"Généré le : {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        ws.cell(row, 1).font = normal_font
        row += 2

        # Statistiques générales
        stats = [
            ("STATISTIQUES GÉNÉRALES", ""),
            ("Enquête", survey),
            ("Nombre total de variables", len(records)),
            ("Variables avec LLM réussi", n_llm_ok),
            ("Variables LLM échoué (heuristique)", n_llm_fail),
            ("Temps de traitement (secondes)", f"{elapsed:.1f}"),
            ("", ""),
            ("QUALITÉ DES DESCRIPTIONS ORIGINALES", ""),
            ("Score 1 (pauvre < 80 chars)", sum(1 for r in records if r.get('qualite') == 1)),
            ("Score 2 (correct 80-300 chars)", sum(1 for r in records if r.get('qualite') == 2)),
            ("Score 3 (riche > 300 chars)", sum(1 for r in records if r.get('qualite') == 3)),
            ("", ""),
            ("DISTRIBUTION DES TYPES", ""),
        ]

        # Comptage des types
        type_counts = {}
        for r in records:
            t = r.get('type_variable', 'autre')
            type_counts[t] = type_counts.get(t, 0) + 1
        for t, cnt in sorted(type_counts.items(), key=lambda x: -x[1]):
            stats.append((f"  • {t}", cnt))

        stats += [
            ("", ""),
            ("DISTRIBUTION DES LANGUES", ""),
        ]
        lang_counts = {}
        for r in records:
            l = r.get('langue_detectee', 'EN')
            lang_counts[l] = lang_counts.get(l, 0) + 1
        for l, cnt in sorted(lang_counts.items(), key=lambda x: -x[1]):
            stats.append((f"  • {l}", cnt))

        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for label, value in stats:
            if label in ("STATISTIQUES GÉNÉRALES", "QUALITÉ DES DESCRIPTIONS ORIGINALES",
                         "DISTRIBUTION DES TYPES", "DISTRIBUTION DES LANGUES"):
                ws.cell(row, 1, label).font = section_font
                ws.cell(row, 1).fill = PatternFill('solid', start_color='D9E1F2')
                ws.cell(row, 2, '').fill = PatternFill('solid', start_color='D9E1F2')
            elif label:
                c1 = ws.cell(row, 1, label)
                c2 = ws.cell(row, 2, value)
                c1.font = normal_font
                c2.font = normal_font
                c1.border = border
                c2.border = border
            row += 1

    def _load_checkpoint(self, survey: str) -> set:
        """Charge les codes déjà traités depuis le fichier de checkpoint."""
        ckpt_path = self.output_dir / f"{survey}_checkpoint.txt"
        if ckpt_path.exists():
            with open(ckpt_path, 'r') as f:
                codes = set(line.strip() for line in f if line.strip())
            log.info(f"   ♻️  Checkpoint: {len(codes)} variables déjà traitées")
            return codes
        return set()

    def _save_checkpoint(self, survey: str, done_codes: set):
        """Sauvegarde les codes traités dans le fichier de checkpoint."""
        ckpt_path = self.output_dir / f"{survey}_checkpoint.txt"
        with open(ckpt_path, 'w') as f:
            for code in sorted(done_codes):
                f.write(code + '\n')

    # ──────────────────────────────────────────────────────────────────────────
    # BOUCLE PRINCIPALE : TRAITEMENT PAR ENQUÊTE
    # ──────────────────────────────────────────────────────────────────────────

    def process_survey(self, survey: str, variables: List[Dict]) -> str:
        """
        Traite toutes les variables d'une enquête et génère le fichier Excel.
        Retourne le chemin du fichier créé.
        """
        log.info(f"\n{'='*70}")
        log.info(f"🔄 TRAITEMENT ENQUÊTE : {survey} ({len(variables)} variables)")
        log.info(f"{'='*70}")

        output_path = self.output_dir / f"{survey}_variables_explained.xlsx"
        start_time  = time.time()

        # Charger le checkpoint (reprise en cas d'interruption)
        done_codes  = self._load_checkpoint(survey)
        records     = []
        n_llm_ok    = 0
        n_llm_fail  = 0

        # Charger les données existantes si le fichier existe déjà
        if output_path.exists() and done_codes:
            try:
                df_existing = pd.read_excel(output_path, sheet_name='Variables')
                records = df_existing.to_dict('records')
                log.info(f"   📂 {len(records)} enregistrements chargés depuis le fichier existant")
            except Exception:
                records = []

        # Traiter chaque variable
        for idx, var in enumerate(variables):
            code = var.get('code', f'UNKNOWN_{idx}')

            # Ignorer si déjà traitée
            if code in done_codes:
                continue

            log.info(f"   [{idx+1}/{len(variables)}] {code}")

            try:
                # ── Heuristique ───────────────────────────────────────────────
                lang    = self._detect_language(
                    f"{var.get('name_en','')}{var.get('description_short','')}"
                )
                vtype   = self._detect_type(
                    code,
                    var.get('name_en', ''),
                    var.get('description_short', '')
                )
                level   = self._detect_level(code, survey,
                                              var.get('description_short', ''))
                quality = self._quality_score(
                    var.get('description_short', ''),
                    var.get('description_long', '')
                )
                heur_short_fr = self._heuristic_short_fr(var)
                heur_long_fr  = self._heuristic_long_fr(var)

                # ── LLM ───────────────────────────────────────────────────────
                llm_results = self._llm_describe(var)
                llm_ok = bool(llm_results.get('long_fr') or llm_results.get('long_orig'))
                if llm_ok:
                    n_llm_ok += 1
                    log.info(f"      ✅ LLM OK")
                else:
                    n_llm_fail += 1
                    log.info(f"      ⚠️  LLM fail → heuristique")

                # ── Enregistrement ────────────────────────────────────────────
                record = {
                    # Identification
                    'code'                    : code,
                    'survey'                  : survey,
                    'category'                : var.get('category', ''),
                    'subcategory'             : var.get('subcategory', ''),
                    'reference_period'        : var.get('reference_period', ''),
                    'langue_detectee'         : lang,
                    'type_variable'           : vtype,
                    'niveau_observation'      : level,
                    'qualite'                 : quality,

                    # Descriptions originales
                    'desc_short_original'     : var.get('description_short', ''),
                    'desc_long_original'      : var.get('description_long', ''),

                    # Heuristique
                    'desc_short_heuristique'  : heur_short_fr,
                    'desc_long_heuristique'   : heur_long_fr,

                    # LLM — langue originale
                    'desc_short_llm_orig'     : llm_results.get('short_orig', heur_short_fr),
                    'desc_long_llm_orig'      : llm_results.get('long_orig',  heur_long_fr),

                    # LLM — français
                    'desc_short_llm_fr'       : llm_results.get('short_fr',   heur_short_fr),
                    'desc_long_llm_fr'        : llm_results.get('long_fr',    heur_long_fr),
                }
                records.append(record)
                done_codes.add(code)

                # ── Checkpoint toutes les N variables ─────────────────────────
                if len(done_codes) % self.checkpoint_every == 0:
                    self._save_checkpoint(survey, done_codes)
                    self._write_excel(output_path, records, survey,
                                      n_llm_ok, n_llm_fail,
                                      time.time() - start_time)
                    log.info(f"   💾 Checkpoint sauvegardé ({len(done_codes)} variables)")

            except Exception as e:
                log.error(f"   ❌ Erreur variable {code}: {e}")
                done_codes.add(code)  # marquer quand même pour éviter boucle infinie

        # Sauvegarde finale
        elapsed = time.time() - start_time
        self._write_excel(output_path, records, survey,
                          n_llm_ok, n_llm_fail, elapsed)
        self._save_checkpoint(survey, done_codes)

        log.info(f"\n   ✅ {survey} terminé : {len(records)} variables traitées")
        log.info(f"   📊 LLM OK: {n_llm_ok} | LLM fail: {n_llm_fail}")
        log.info(f"   ⏱️  Temps: {elapsed:.1f}s")
        log.info(f"   📁 Fichier: {output_path}")

        return str(output_path)

    def _write_excel(self, path: Path, records: List[Dict], survey: str,
                     n_llm_ok: int, n_llm_fail: int, elapsed: float):
        """Écrit (ou réécrit) le fichier Excel avec toutes les données."""
        wb = Workbook()

        # ── Feuille principale : Variables ────────────────────────────────────
        ws = wb.active
        ws.title = "Variables"

        headers = [
            'Code', 'Enquête', 'Catégorie', 'Sous-catégorie',
            'Période réf.', 'Langue', 'Type variable',
            'Niveau obs.', 'Qualité orig.',
            'Description courte originale',
            'Description longue originale',
            'Description courte (heuristique FR)',
            'Description longue (heuristique FR)',
            'Description courte LLM (langue orig.)',
            'Description longue LLM (langue orig.)',
            'Description courte LLM (FR)',
            'Description longue LLM (FR)',
        ]

        col_map = [
            'code', 'survey', 'category', 'subcategory',
            'reference_period', 'langue_detectee', 'type_variable',
            'niveau_observation', 'qualite',
            'desc_short_original', 'desc_long_original',
            'desc_short_heuristique', 'desc_long_heuristique',
            'desc_short_llm_orig', 'desc_long_llm_orig',
            'desc_short_llm_fr', 'desc_long_llm_fr',
        ]

        # En-têtes
        for col, header in enumerate(headers, 1):
            ws.cell(1, col, header)

        # Données
        for row_idx, record in enumerate(records, 2):
            for col_idx, key in enumerate(col_map, 1):
                ws.cell(row_idx, col_idx, record.get(key, ''))

        self._style_excel(ws, len(records))

        # ── Feuille Résumé ────────────────────────────────────────────────────
        self._add_summary_sheet(wb, survey, records, n_llm_ok, n_llm_fail, elapsed)

        wb.save(path)

    def process_all_surveys(self, variables_by_survey: Dict[str, List[Dict]]):
        """Lance la boucle pour toutes les enquêtes."""
        print("\n" + "="*70)
        print("🚀 BOUCLE D'EXPLICATION — TOUTES LES ENQUÊTES")
        print("="*70)

        total_start = time.time()
        results = {}

        for survey, variables in variables_by_survey.items():
            if not variables:
                log.warning(f"⚠️  Enquête {survey} : aucune variable trouvée")
                continue
            path = self.process_survey(survey, variables)
            results[survey] = path

        total_elapsed = time.time() - total_start
        print(f"\n{'='*70}")
        print(f"✅ TOUTES LES ENQUÊTES TRAITÉES")
        print(f"   ⏱️  Temps total : {total_elapsed:.1f}s")
        print(f"\n   Fichiers générés :")
        for survey, path in results.items():
            print(f"      • {path}")
        print("="*70)

        return results


# ══════════════════════════════════════════════════════════════════════════════
# DÉMONSTRATION
# ══════════════════════════════════════════════════════════════════════════════

def demo():
    print("="*80)
    print("DÉMONSTRATION SYSTÈME RAG (E5-LARGE + COSINE + EXPLAINER)")
    print("="*80)

    chroma_path    = Path('../data/embeddings/chroma_db')
    variables_path = Path('../data/unified/unified_variables.json')
    output_dir     = Path('../data/explained')

    if not chroma_path.exists():
        chroma_path    = Path('data/embeddings/chroma_db')
        variables_path = Path('data/unified/unified_variables.json')
        output_dir     = Path('data/explained')

    # ── Initialisation RAG ────────────────────────────────────────────────────
    print("🔧 Initialisation du système RAG...")
    rag = MultiSurveyRAG(chroma_path, variables_path)

    print(f"\n   ✅ Collections chargées:")
    for name, coll in rag.collections.items():
        print(f"      • {name}: {coll.count()} variables")

    # ── TEST RECHERCHE SIMPLE ─────────────────────────────────────────────────
    print("\n" + "="*80)
    print("TEST : Recherche sémantique simple (sans pondération)")
    print("="*80)

    questions = [
        "Heures travaillées par semaine",
        "Variables sur le patrimoine immobilier",
        "Déclaration d'impôt revenus cadastraux",
    ]
    for question in questions:
        print(f"\n❓ {question}")
        print("-"*80)
        rag.display_results(rag.search_by_question(question, top_k=5))

    # ── TEST RECHERCHE PAR CODE ───────────────────────────────────────────────
    print("\n" + "="*80)
    print("TEST : Code → Description")
    print("="*80)
    for code in ['HY040N', 'A0270', 'abo_h_fm']:
        print(f"\n🔍 Code: {code}")
        result = rag.search_by_code(code)
        if result['found']:
            print(f"   ✅ {result['survey']} | {result['name']}")
        else:
            print(f"   ❌ {result['message']}")

    # ── BOUCLE D'EXPLICATION ─────────────────────────────────────────────────
    print("\n" + "="*80)
    print("BOUCLE D'EXPLICATION DES VARIABLES")
    print("="*80)

    if not rag.variables_by_survey:
        print("⚠️  Aucune variable chargée — vérifier variables_path")
        return

    explainer = VariableExplainer(
        output_dir     = output_dir,
        ollama_model   = 'llama3.2',
        checkpoint_every = 50
    )

    # Lancer sur toutes les enquêtes
    explainer.process_all_surveys(rag.variables_by_survey)

    print("\n" + "="*80)
    print("✅ DÉMONSTRATION TERMINÉE")
    print("="*80)
    print("\n🚀 Prochaine étape: python step4_llm_wrapper.py")


if __name__ == '__main__':
    try:
        demo()
    except Exception as e:
        print(f"\n❌ ERREUR: {e}")
        import traceback
        traceback.print_exc()
